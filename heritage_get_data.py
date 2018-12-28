from selenium import webdriver
import time
import openpyxl
import copy
import requests
import os
driver = webdriver.Chrome("C:\\ProgramData\\Anaconda3\\selenium\\webdriver\\chromedriver_win32\\chromedriver.exe")

l = []

driver.get('https://coins.ha.com/c/search-results.zx?N=790+231+51&ic10=ArchiveTab-071515')
time.sleep(1)

book = openpyxl.load_workbook('C:\\Users\\深澤裕士\\Desktop\\heritage.xlsx')
sheet = book['Sheet1']
sheet.cell(row=1,column=1).value = 'Number'
sheet.cell(row=1,column=2).value = 'Grade'
sheet.cell(row=1,column=3).value = 'Title'
sheet.cell(row=1,column=4).value = 'Description'
s = 1
#site = "https://coins.ha.com/c/search-results.zx?N=790+231+51&No=50&erpp=50"#.format(num=s)
site = "https://coins.ha.com/c/search-results.zx?Ne=109&N=790+231+51+1994&ic4=Refine-Highlights-360View-102615"
driver.get(site)
#//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[1]/input
#//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[1]/input
cat_n = 1
lot_n = 1
limit = 0
category = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[{n}]/input'.format(n=cat_n))#.format(n=cat_n))
grade = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[{n}]'.format(n=cat_n)).text#.format(n=cat_n)).text
grade = grade.split(' (')[0]
category.click()
time.sleep(1)

def heritage(cat_n, lot_n):
    limit = 0
    #site = "https://coins.ha.com/c/search-results.zx?Ne=109&N=790+231+51+1994&ic4=Refine-Highlights-360View-102615"
    site = "https://coins.ha.com/c/search-results.zx?N=790+231+51&ic10=ArchiveTab-071515"
    driver.get(site)
    time.sleep(1)
    category = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[{n}]/input'.format(n=cat_n))#.format(n=cat_n))
    grade = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[{n}]'.format(n=cat_n)).text#.format(n=cat_n)).text
    grade = grade.split(' (')[0]
    category.click()
    time.sleep(1)

    while(True):
        try:
            print('limit : ', limit)
            print('cat_n : ', cat_n)
            print('grade : ', grade)
            if limit == 0:
                os.mkdir('E:\\機械学習用\\'+grade)

            lots = driver.find_elements_by_css_selector('a.item-title')
            if len(lots)>50:
                lots = lots[:50]
            for lot in range(len(lots)):
                print('cat_n : ',cat_n)
                print('lot:', lot+1, 'len:', len(lots), 'selector:', lots[lot])

                # 1つのcategoryにつき1000個分を上限とする
                if limit >= 1000:
                    return lot_n

                lots_kari = driver.find_elements_by_css_selector('a.item-title')
                lots_kari[lot].click()
                time.sleep(1)
                try:
                    title = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[5]/div[1]/div[3]/div/div/h1').text
                    text = driver.find_element_by_xpath('//*[@id="auction-description"]/span').text
                    grade = grade
                    sheet.cell(row=lot_n+1,column=1).value = lot_n
                    sheet.cell(row=lot_n+1,column=2).value = grade
                    sheet.cell(row=lot_n+1,column=3).value = title
                    sheet.cell(row=lot_n+1,column=4).value = text
                    book.save('C:\\Users\\深澤裕士\\Desktop\\heritage.xlsx')
                    img_n = 0
                    while(True):
                        try:
                            if img_n != 0:
                                selector = '//*[@id="image-nav-{n}"]/img'.format(n=img_n)
                                s = driver.find_element_by_xpath(selector)
                                s.click()
                                time.sleep(1)
                            img_selector = '//*[@id="page-content"]/div/div[5]/div[1]/div[3]/div/div/div[1]/div[1]/ul/div/div/li[{n}]/span[1]/img'.format(n=str(img_n+2))
                            src = driver.find_element_by_xpath(img_selector).get_attribute('src')
                            res = requests.get(src)
                            time.sleep(1)
                            file_name = 'img' + '_' + str(lot_n) + '_' + str(img_n+1) + '.jpg'
                            with open('E:\\機械学習用\\' + grade+ '\\' + file_name, "wb") as img:
                                img.write(res.content)
                            img_n += 1
                            print(file_name)
                        except:
                            break
                    driver.back()
                    time.sleep(1)
                    limit += 1
                    lot_n += 1
                except:
                    book.save('C:\\Users\\深澤裕士\\Desktop\\heritage.xlsx')
                    print('end')
                    break
            try:
                next = driver.find_element_by_css_selector('a.pageindex.direction.icon-right-triangle')
                #print('next2', next)
                next.click()
                #print('next3')
                time.sleep(1)
                #print('next4')
                sheet.cell(row=lot_n+1,column=5).value = '次ページへ'
                print('----------------------------------Next-----------------------------------')
            except:
                sheet.cell(row=lot_n+1,column=6).value = '最後まで見ました'
                print('------------------------This category is Finish!-------------------------')
                return lot_n
        except:
            end = input('finish or continue : ')
            if end == 'finish':
                print('Finish!')
                return lot_n
            else:
                continue

for i in range(34):
    print('lot_n : ', lot_n)
    num = copy.copy(lot_n)
    print('num : ', num)
    lot_n = heritage(i+1, num)
