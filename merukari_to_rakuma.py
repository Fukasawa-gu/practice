# -*- coding: utf-8 -*-
"""
Created on Wed Jun 13 14:57:26 2018

@author: Fukasawa-gu
"""

#メルカリのデータをスクレイピング
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementNotVisibleException
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import datetime as dt
import re
import time
import datetime
import xlwt
import os
import urllib.request
import xlrd
import types
import requests
import pandas as pd

# Excelのシート作成
#book = xlwt.Workbook()
#sheet1=book.add_sheet('sheet1')
wb = Workbook()
book = wb['Sheet']

book['b1'].value = 'タイトル'
book['c1'].value = '価格'
book['d1'].value = '詳細'
book['e1'].value = '商品の状態'
book['f1'].value = '画像1'
book['g1'].value = '画像2'
book['h1'].value = '画像3'
book['i1'].value = '画像4'

# chroを定義
try:
    chro = webdriver.Chrome()
except:
    chro = webdriver.Chrome('C:/selenium/chromedriver')

def searchLink(linkString):
    elem_last_btn=chro.find_element_by_link_text(linkString)
    elem_last_btn.click()
    time.sleep(1)

def xpath_click(xpath):
    input = chro.find_element_by_xpath(xpath)
    input.click()
    time.sleep(1)

#メルカリでログイン
chro.get('https://www.mercari.com/jp/')

chro.find_element_by_link_text('ログイン').click()
time.sleep(1)

input = chro.find_element_by_xpath('/html/body/div[1]/main/div/form/div/div[1]/input')
time.sleep(1)

input.send_keys('XXXXX@yahoo.co.jp')
time.sleep(1)

input = chro.find_element_by_xpath('/html/body/div[1]/main/div/form/div/div[2]/input')

time.sleep(1)

input.send_keys('XXXXX')

time.sleep(60)# 私はロボットではありませんを回避する

input = chro.find_element_by_xpath('/html/body/div[1]/main/div/form/div/button')
input.click()
time.sleep(1)


"""page_dic = {}"""
try:
    os.mkdir('C:/Users/Fukasawa-gu/Desktop/merukari_data')
except:
    None

page = 1
try:
    k = 0
    while(True):
        if page == 1:
            url = 'https://www.mercari.com/jp/mypage/listings/listing/'
        else:
            url= chro.current_url
        print(page, url)
        for i in range(1,51):
            chro.get(url)
            time.sleep(1)
            """if i == 51:
                print(i)
                break"""

            open = chro.find_element_by_xpath('//*[@id="mypage-tab-transaction-now"]/li[{}]/a/div/div[2]/div'.format(i)).get_attribute('innerHTML')
            time.sleep(1)
            print(open)
            if '公開停止中' in open:
                print(open + 'なので飛ばします。')
                continue

            title = chro.find_element_by_xpath('//*[@id="mypage-tab-transaction-now"]/li[{}]/a/div/div[1]'.format(i)).get_attribute('innerHTML')
            time.sleep(1)
            print(title)
            # 指定の文字がタイトルに含まれていた時は以下のコメントアウトを外す
            """s = ''
            if s in title:
                print(s + 'が含まれているので飛ばします。')
                continue"""

            # 商品ページに移動
            xpath_click('//*[@id="mypage-tab-transaction-now"]/li[{}]/a/div/div[2]/div'.format(i))

            # タイトル、詳細、値段、発送元、商品の状態
            title = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/h2').get_attribute('innerHTML')
            description = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[3]').get_attribute('innerHTML')
            price = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[2]/span[1]').get_attribute('innerHTML')
            price = price.replace(',','')
            price = price.replace('¥','')
            price = price.replace(' ','')
            condition = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[1]/table/tbody/tr[4]/td').get_attribute('innerHTML')

            print('-----------------------------------')
            print(chro.current_url)
            print('-----------------------------------')
            print(title)
            print('-----------------------------------')
            print(description)
            print('-----------------------------------')
            print(price)
            print('-----------------------------------')
            print(condition)

            gazos = {}
            filenames = []
            for j in range(1,5):
                try:
                    xpath_click('/html/body/div/main/div[1]/section/div[1]/div/div/div[3]/div[%s]'%j)
                    gazos[j] = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[1]/div/div/div[1]/div/div[%s]/div/img'%j).get_attribute('src')
                    print('gazos[j]',gazos[j])
                    #res = requests.get(gazos[j])
                    print('open',page,i,j)
                    path = 'C:/Users/Fukasawa-gu/Desktop/merukari_data/img{0}_{1}_{2}.jpg'.format(page,i,j)
                    filenames.append('img{0}_{1}_{2}.jpg'.format(page,i,j))
                    print(path)
                    urllib.request.urlretrieve(gazos[j], path)

                except:
                    print('画像が4枚未満')
                    pass

            print(gazos)

            book = wb['Sheet']
            book['b'+str(i+k+1)].value = title
            book['c'+str(i+k+1)].value = price
            book['d'+str(i+k+1)].value = description
            book['e'+str(i+k+1)].value = condition
            try:
                book['f'+str(i+k+1)].value = filenames[0]
                book['g'+str(i+k+1)].value = filenames[1]
                book['h'+str(i+k+1)].value = filenames[2]
                book['i'+str(i+k+1)].value = filenames[3]
            except:
                pass
            wb.save('C:/Users/Fukasawa-gu/Desktop/merukari_data/merukari_data.xlsx')

            chro.back()
            time.sleep(1)
        k += 50
        page += 1
        xpath_click('/html/body/div/main/div[1]/ul/li[2]/a')

except Exception as e:
    print(e)
    print('最終ページまで行きつきました。')

print('メルカリのデータ取得を終えました。')

book = pd.read_excel("C:/Users/Fukasawa-gu/Desktop/merukari_data/merukari_data.xlsx", index_col=0)

def searchLink(linkString):
    elem_last_btn=chro.find_element_by_link_text(linkString)
    elem_last_btn.click()
    time.sleep(1)

def xpath_click(xpath):
    input = chro.find_element_by_xpath(xpath)
    input.click()
    time.sleep(1)

chro = webdriver.Chrome()
name = "XXXXX@yahoo.co.jp"
passwd = "XXXXX"
chro.get('https://fril.jp/')
time.sleep(1)

searchLink("ログイン")

input = chro.find_element_by_id('email')
input.send_keys("{}".format(name))
time.sleep(1)


input = chro.find_element_by_id('password')
input.send_keys("{}".format(passwd))
time.sleep(1)

xpath_click('//*[@id="new_user"]/div[3]/input')

try:
    for i in range(len(book)):
        chro.get('https://fril.jp/item/new')
        try:
            title = book.iloc[i, 0]
            price = book.iloc[i, 1]
            description = book.iloc[i, 2].replace('<br>','\n')
            condition = book.iloc[i, 3]
            photo1 = book.iloc[i, 4]
            photo2 = book.iloc[i, 5]
            photo3 = book.iloc[i, 6]
            photo4 = book.iloc[i, 7]
            photos = [photo1, photo2, photo3, photo4]
            print(title)
            print(price)
            print(description)
            print(condition)
            print(photo1)
            print(photo2)
            print(photo3)
            print(photo4)

            dirname = 'C:/Users/Fukasawa-gu/Desktop/merukari_data/'
            for j in range(4):
                try:
                    dir_photo = dirname + photos[j]
                    upload_photo = chro.find_element_by_xpath('//*[@id="files"]/div[{}]/div[2]/input'.format(j+1))
                    time.sleep(1)
                    upload_photo.send_keys(dir_photo)
                    time.sleep(1)
                except Exception as e:
                    print(e)
                    pass

            xpath_click('//*[@id="category_name"]')
            xpath_click('//*[@id="select-category"]/div/div/div[2]/div/div[14]/a')
            xpath_click('//*[@id="menu_13"]/div[2]/a')
            xpath_click('//*[@id="menu_13_1"]/a[2]')

            for j in range(1,7):
                status = chro.find_element_by_xpath('//*[@id="status"]/option[{}]'.format(j)).get_attribute('innerHTML')
                if status == condition:
                    xpath_click('//*[@id="status"]/option[{}]'.format(j))

            xpath_click('//*[@id="delivery_date"]')
            xpath_click('//*[@id="delivery_date"]/option[2]')

            xpath_click('//*[@id="delivery_area"]')
            xpath_click('//*[@id="delivery_area"]/option[13]')

            input_title = chro.find_element_by_xpath('//*[@id="name"]')
            time.sleep(1)
            input_title.send_keys(title)
            time.sleep(1)

            input_description = chro.find_element_by_xpath('//*[@id="detail"]')
            time.sleep(1)
            input_description.send_keys(description)
            time.sleep(1)

            input_price = chro.find_element_by_xpath('//*[@id="sell_price"]')
            time.sleep(1)
            input_price.send_keys(int(price))
            time.sleep(1)

            xpath_click('//*[@id="confirm"]')
            time.sleep(3)
            xpath_click('//*[@id="submit"]')
            time.sleep(3)
        except:
            continue
except Exception as e:
    print(e)
    print('ラクマに出品を完了しました。')