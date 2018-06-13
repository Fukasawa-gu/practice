#メルカリのデータをスクレイピング
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementNotVisibleException
from selenium.common.exceptions import WebDriverException
import re
import time
import datetime
#import requests
import xlwt
import os
import urllib.request
from openpyxl import Workbook, load_workbook

# Excelのシート作成
#book = xlwt.Workbook()
#sheet1=book.add_sheet('sheet1')
wb = Workbook()
book = wb['Sheet']

book['b1'].value = 'タイトル'
book['c1'].value = '価格'
book['d1'].value = '詳細'
book['e1'].value = '商品の状態'
book['f1'].value = '画像1'
book['g1'].value = '画像2'
book['h1'].value = '画像3'
book['i1'].value = '画像4'

# chroを定義
try:
    chro = webdriver.Chrome()
except:
    chro = webdriver.Chrome('C:/selenium/chromedriver')

def searchLink(linkString):
    elem_last_btn=chro.find_element_by_link_text(linkString)
    elem_last_btn.click()
    time.sleep(1)

def xpath_click(xpath):
    input = chro.find_element_by_xpath(xpath)
    input.click()
    time.sleep(1)

#メルカリでログイン
chro.get('https://www.mercari.com/jp/')

chro.find_element_by_link_text('ログイン').click()
time.sleep(1)

input = chro.find_element_by_xpath('/html/body/div[1]/main/div/form/div/div[1]/input')
time.sleep(1)

input.send_keys('XXXXX@yahoo.co.jp')
time.sleep(1)

input = chro.find_element_by_xpath('/html/body/div[1]/main/div/form/div/div[2]/input')

time.sleep(1)

input.send_keys('XXXXX')

time.sleep(60)# 私はロボットではありませんを回避する

input = chro.find_element_by_xpath('/html/body/div[1]/main/div/form/div/button')
input.click()
time.sleep(1)


"""page_dic = {}"""
try:
    os.mkdir('C:/Users/Fukasawa-gu/Desktop/merukari_data')
except:
    None

page = 1
try:
    k = 0
    while(True):
        if page == 1:
            url = 'https://www.mercari.com/jp/mypage/listings/listing/'
        else:
            url= chro.current_url
        print(page, url)
        for i in range(1,51):
            chro.get(url)
            time.sleep(1)
            """if i == 51:
                print(i)
                break"""

            open = chro.find_element_by_xpath('//*[@id="mypage-tab-transaction-now"]/li[{}]/a/div/div[2]/div'.format(i)).get_attribute('innerHTML')
            time.sleep(1)
            print(open)
            if '公開停止中' in open:
                print(open + 'なので飛ばします。')
                continue

            title = chro.find_element_by_xpath('//*[@id="mypage-tab-transaction-now"]/li[{}]/a/div/div[1]'.format(i)).get_attribute('innerHTML')
            time.sleep(1)
            print(title)
            # 指定の文字がタイトルに含まれていた時は以下のコメントアウトを外す
            """s = ''
            if s in title:
                print(s + 'が含まれているので飛ばします。')
                continue"""

            # 商品ページに移動
            xpath_click('//*[@id="mypage-tab-transaction-now"]/li[{}]/a/div/div[2]/div'.format(i))

            # タイトル、詳細、値段、発送元、商品の状態
            title = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/h2').get_attribute('innerHTML')
            description = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[3]').get_attribute('innerHTML')
            price = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[2]/span[1]').get_attribute('innerHTML')
            price = price.replace(',','')
            price = price.replace('¥','')
            price = price.replace(' ','')
            condition = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[1]/table/tbody/tr[4]/td').get_attribute('innerHTML')

            print('-----------------------------------')
            print(chro.current_url)
            print('-----------------------------------')
            print(title)
            print('-----------------------------------')
            print(description)
            print('-----------------------------------')
            print(price)
            print('-----------------------------------')
            print(condition)

            gazos = {}
            filenames = []
            for j in range(1,5):
                try:
                    xpath_click('/html/body/div/main/div[1]/section/div[1]/div/div/div[3]/div[%s]'%j)
                    gazos[j] = chro.find_element_by_xpath('/html/body/div/main/div[1]/section/div[1]/div/div/div[1]/div/div[%s]/div/img'%j).get_attribute('src')
                    print('gazos[j]',gazos[j])
                    #res = requests.get(gazos[j])
                    print('open',page,i,j)
                    path = 'C:/Users/Fukasawa-gu/Desktop/merukari_data/img{0}_{1}_{2}.jpg'.format(page,i,j)
                    filenames.append('img{0}_{1}_{2}.jpg'.format(page,i,j))
                    print(path)
                    urllib.request.urlretrieve(gazos[j], path)

                except:
                    print('画像が4枚未満')
                    pass

            print(gazos)

            book = wb['Sheet']
            book['b'+str(i+k+1)].value = title
            book['c'+str(i+k+1)].value = price
            book['d'+str(i+k+1)].value = description
            book['e'+str(i+k+1)].value = condition
            try:
                book['f'+str(i+k+1)].value = filenames[0]
                book['g'+str(i+k+1)].value = filenames[1]
                book['h'+str(i+k+1)].value = filenames[2]
                book['i'+str(i+k+1)].value = filenames[3]
            except:
                pass
            wb.save('C:/Users/Fukasawa-gu/Desktop/merukari_data/merukari_data.xlsx')

            chro.back()
            time.sleep(1)
        k += 50
        page += 1
        xpath_click('/html/body/div/main/div[1]/ul/li[2]/a')

except Exception as e:
    print(e)
    print('最終ページまで行きつきました。')

print('end')
