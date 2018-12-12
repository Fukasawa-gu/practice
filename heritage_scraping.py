from selenium import webdriver
import time
import openpyxl
import copy
driver = webdriver.Chrome("C:\\ProgramData\\Anaconda3\\selenium\\webdriver\\chromedriver_win32\\chromedriver.exe")

l = []

driver.get('https://coins.ha.com/c/search-results.zx?N=790+231+51&ic10=ArchiveTab-071515')
time.sleep(1)

book = openpyxl.load_workbook('C:\\Users\\深澤裕士\\Desktop\\heritage.xlsx')
sheet = book['Sheet1']
s = 1
site = "https://coins.ha.com/c/search-results.zx?N=790+231+51&No=50&erpp=50".format(num=s)
driver.get(site)
#//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[1]/input
#//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[1]/input
cat_n = 1
lot_n = 1
while True:
    print(cat_n)
    book.save('C:\\Users\\深澤裕士\\Desktop\\chromedriver_win32\\heritage.xlsx')
    #//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[4]/a[1]
    #//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[4]/a[2]
    category = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[10]/a[{n}]/input'.format(n=cat_n))
    grade = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[4]/div[1]/div[2]/div/div/div[4]/a[{n}]'.format(n=cat_n)).text
    category.click()
    time.sleep(1)
    lots = driver.find_elements_by_css_selector('a.item-title')
    print(len(lots))
    if len(lots)>50:
        lots = lots[:50]
    print(len(lots))
    for lot in lots:
        lot.click()
        time.sleep(1)
        try:
            title = driver.find_element_by_xpath('//*[@id="page-content"]/div/div[5]/div[1]/div[3]/div/div/h1').text
            text = driver.find_element_by_xpath('//*[@id="auction-description"]/span').text
            grade = grade
            print(grade)
            print(title)
            print(text)
            sheet.cell(row=lot_n,column=1).value = grade
            sheet.cell(row=lot_n,column=2).value = title
            sheet.cell(row=lot_n,column=3).value = text
            book.save('C:\\Users\\深澤裕士\\Desktop\\heritage.xlsx')
            img_n = 0
            while(True):
                try:
                    selector = '//*[@id="image-nav-{n}"]/img'.format(n=img_n)
                    selector.click()
                    time.sleep(1)
                    img_selector = '//*[@id="page-content"]/div/div[5]/div[1]/div[3]/div/div/div[1]/div[1]/ul/div/div/li[{n}]/span[1]/img'.format(n=img_n+2)
                    src = driver.find_element(img_selector).attribute('src')
                    request = urllib.request.urlopen(url)
                    file_name = 'img' + '_' + str(lot_n) + '_' + str(img_n) + '.jpg'
                    f = open('C:\\Users\\深澤裕士\\Desktop\\' + file_name, "wb")
                    f.write(request.read())
                    f.close()
                    img_n += 1
                except:
                    break
        except:
            book.save('C:\\Users\\深澤裕士\\Desktop\\heritage.xlsx')
            print('end')
            break
    break
